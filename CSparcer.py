import os
import subprocess
from urllib.request import urlopen
from bs4 import BeautifulSoup
import json
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time
import psutil


def get_category_type(item_type):
    if item_type == "knives":
        return "tag_CSGO_Type_Knife"
    elif item_type == "gloves":
        return "tag_Type_Hands&appid=730"
    elif item_type == "stickers":
        return "tag_TeamLogo&appid=730#p1_quantity_desc"
    else:
        return "tag_CSGO_Type_Knife&category_730_Type%5B%5D=tag_Type_Hands&appid=730"

items_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "items.txt")
if os.path.exists(items_file_path):
    print("items.txt file found. Parsing links from it")
    subprocess.run(["python", "CSparcer2.py"])
else:
    num_items = int(input("Enter the number of links to parse: "))
    item_type = input("Enter the item type for parsing (knives, gloves, both (knives or gloves) or stickers): ").lower()
    category_type = get_category_type(item_type)
    if item_type == "stickers":
        link = f'https://steamcommunity.com/market/search/render/?query=&start=0&count={num_items}&search_descriptions=0&sort_column=quantity&sort_dir=desc&appid=730&category_730_ItemSet%5B%5D=any&category_730_ProPlayer%5B%5D=any&category_730_StickerCapsule%5B%5D=any&category_730_TournamentTeam%5B%5D=any&category_730_Weapon%5B%5D=any&category_730_StickerCategory%5B%5D=tag_PlayerSignature&category_730_StickerCategory%5B%5D=tag_TeamLogo&category_730_StickerCategory%5B%5D=tag_Tournament&appid=730'
    else:
        link = f'https://steamcommunity.com/market/search/render/?query=&start=0&count={num_items}&search_descriptions=0&sort_column=price&sort_dir=asc&appid=730&category_730_ItemSet%5B%5D=any&category_730_ProPlayer%5B%5D=any&category_730_StickerCapsule%5B%5D=any&category_730_TournamentTeam%5B%5D=any&category_730_Weapon%5B%5D=any&category_730_Type%5B%5D={category_type}'
    page = urlopen(link)
    data = json.loads(page.read().decode())
    html = data['results_html']
    bs_page = BeautifulSoup(html, features="html.parser")
    objects = bs_page.findAll(class_="market_listing_row_link")

    urls = []

    for g in objects:
     link = g["href"]
     price = g.find('span', {'data-price': True}).text
     urls.append(link)
     print(f"{price} | {link}")
    

if not urls:
    print("Steam API servers are currently overloaded (or you just got a ~15min ban for requesting too much)")

options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument('--window-size=1280x1696')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--remote-debugging-port=9222')

chrome_driver_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chromedriver.exe")
service = Service(executable_path=chrome_driver_path)

driver = webdriver.Chrome(service=service, options=options)


def extract_number(text):
    return float(''.join(filter(str.isdigit, text))) / 100

data = []
request_counter = 0

for url in urls:
    print(f"Processing URL: {url}")
    driver.get(url)
    time.sleep(2) 
    browser_logs = driver.get_log("browser")
    print("Browser Console Logs:", browser_logs)
    wait = WebDriverWait(driver, 20)

    try:
        span_elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "span.market_commodity_orders_header_promote")))
        if len(span_elements) == 2:
            buy_requests = span_elements[0].text
            order_price = span_elements[1].text
            order_price_num = extract_number(order_price)
            sell_price = order_price_num
            starting_price = None
            for price, link in zip(objects, urls):
                if link == url:
                    starting_price = 0.87*extract_number(price.find('span', {'data-price': True}).text)
                    break
            if starting_price is not None:
                profit =  starting_price - sell_price
                print(f"Buy Requests: {buy_requests}, Selling Price: {starting_price:.2f}, Order Price: {sell_price:.2f}, Profit: {profit:.2f}, Link: {url}")
                data.append((buy_requests, sell_price, starting_price, profit))
            else:
                print("Failed to find the selling price.")
        else:
            print("Failed to find two span elements with the specified class.")

    except Exception as e:
     print(f"An error occurred while parsing price: {type(e)} - {e}")

    request_counter += 1
    if request_counter % 24 == 0:
        print("Request limit reached. Waiting for the opportunity to continue...")
        for remaining_time in range(270, 0, -5):
            time.sleep(5)
            print(f"Remaining time: {remaining_time}sec...")




driver.quit()

def close_excel():
    for process in psutil.process_iter():
        if process.name() == 'EXCEL.EXE':
            process.terminate()
            print("Excel has been closed")
            time.sleep(1)  

close_excel()


filename = "parser.xlsx"


if os.path.isfile(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active

    
    headers = ["Buy Requests", "Selling Price", "Order Price", "Profit", "Link"]
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.cell.get_column_letter(col_num)
        ws[f"{col_letter}1"] = header


first_empty_row = len(ws["A"]) + 1
for row_num, (buy_requests, sell_price, starting_price, profit) in enumerate(data, start=first_empty_row):
    ws.cell(row=row_num, column=1, value=buy_requests)
    ws.cell(row=row_num, column=2, value=sell_price)
    ws.cell(row=row_num, column=3, value=starting_price)
    ws.cell(row=row_num, column=4, value=profit)
    ws.cell(row=row_num, column=5, value=urls[row_num - first_empty_row])


wb.save(filename)


print(data)
print('Excel updated.')
