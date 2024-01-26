import os
import psutil
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup

def close_excel():
    for process in psutil.process_iter():
        if process.name() == 'EXCEL.EXE':
            process.terminate()
            print("Excel has been closed")
            time.sleep(1)  

close_excel()

filename = "parcedstickers.xlsx"

if os.path.isfile(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ["Name", "Buy Requests", "Selling Price", "Profit", "Link"]
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.cell.get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument('--window-size=1280x1696')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--remote-debugging-port=9222')

current_dir = os.path.dirname(os.path.abspath(__file__))
items_file_path = os.path.join(current_dir, 'items.txt')
chrome_driver_path = os.path.join(current_dir, 'chromedriver.exe')
webdriver_service = Service(executable_path=chrome_driver_path)
driver = webdriver.Chrome(service=webdriver_service, options=options)

with open(items_file_path, 'r') as file:
    urls = file.readlines()

print(f"Read {len(urls)} URLs from {items_file_path}")

data = []
request_counter = 0
for url in urls:
    driver.get(url.strip())
    time.sleep(2)
    WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'market_commodity_orders_header_promote')))

    soup = BeautifulSoup(driver.page_source, 'html.parser')

    item_name = soup.find('span', {'class': 'market_listing_item_name'}).text

    spans = soup.find_all('span', {'class': 'market_commodity_orders_header_promote'})

    sell_price = float(spans[1].text.strip().replace('$', ''))
    buy_request_price = float(spans[3].text.strip().replace('$', ''))

    profit = sell_price * 0.87 - buy_request_price

    if 0 < profit < 0.01:
        profit = 0

    print(f"Name: {item_name} | Buyrequest price: {buy_request_price} | Sellprice: {sell_price} | Profit: {profit}")

    data.append((item_name, buy_request_price, sell_price, profit, url.strip()))

    request_counter += 1
    if request_counter % 24 == 0:
        print("Request limit reached. Waiting for the opportunity to continue...")
        for remaining_time in range(270, 0, -5):
            time.sleep(5)
            print(f"Remaining time: {remaining_time}sec...")

driver.quit()

first_empty_row = len(ws["A"]) + 1
for row_num, (name, buy_requests, sell_price, profit, link) in enumerate(data, start=first_empty_row):
    ws.cell(row=row_num, column=1, value=name)
    ws.cell(row=row_num, column=2, value=buy_requests)
    ws.cell(row=row_num, column=3, value=sell_price)
    ws.cell(row=row_num, column=4, value=profit)
    ws.cell(row=row_num, column=5, value=link)

wb.save(filename)

print(data)
print('Excel updated.')